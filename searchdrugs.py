import time
import re
import openpyxl

def search_drugs():
    found =False
    search_results = []
    print("****************************************************************")
    print("***********************Search Drugs*****************************")
    print("****************************************************************")
    drug_name=input("Enter the drug name to search : ")
    print(f'Searching for {drug_name} in the inventory...')
    #Open the inventory file
    medicine_inventory = openpyxl.load_workbook("medicine_inventory.xlsx")
    mi = medicine_inventory.active

    for i in range(1,mi.max_row+1): 
        for j in range(1,2):
            med_id = mi.cell(row=i, column=1)                            
            med_name = mi.cell(row=i, column=2)
            med_qty = mi.cell(row=i, column=4) 
            match_drug = re.match('.*'+drug_name.lower()+'.*',med_name.value.lower())
            if (match_drug):                
                found = True
                print(f" Medicine Id : {med_id.value}",end=" ")
                print(f" Medicine Name : {med_name.value}",end=" ")
                print(f" Quantity Available : {med_qty.value}",end="\n")
                search_results.append(med_id.value)
    if found is False:
        print("Medicine Not Found")        
    else:
        print(f'search results {search_results}')