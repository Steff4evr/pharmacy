import openpyxl

def invoice_update(cart, invoice_number):
    invoices = openpyxl.load_workbook("invoices.xlsx")
    update_invoices = invoices.active
    print(f"invoice number : {invoice_number}")
    print(f"cart : {cart}")
    j= update_invoices.max_row
    for i in cart:
        j += 1
        print("invoice update2")
        print(i["med_id"], i["med_name"], i["med_qty"], i["med_price"])
        update_invoices.cell(row=j, column=1).value = invoice_number
        update_invoices.cell(row=j, column=2).value = i["med_id"]
        update_invoices.cell(row=j, column=3).value = i["med_name"]
        update_invoices.cell(row=j, column=4).value = i["med_qty"]
        update_invoices.cell(row=j, column=5).value = i["med_price"]*i["med_qty"]
    invoices.save("invoices.xlsx")
    print("Updated the invoices report .. ")
