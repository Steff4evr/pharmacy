import time
from datetime import datetime
import uuid
import openpyxl

def billing_invoice_generation(cart):
    total = 0
    while True:
        print("\n")
        print("****************************************************************")
        print("********************Billing & Invoice**********************")
        print("****************************************************************")
        print(f'Medicines in cart: {cart}')
        try:
            billing_user_input = input("Do you want to continue with the billing ? (y/n)")
            if billing_user_input.lower() == "y":
                #Creating the invoice 
                # create a company name and information

                company_name = 'Steff''s Pharmacy.'
                company_address = '007 James Bond St.'
                company_city = 'Melbourne'                               
                # declare ending message
                message = 'Thanks for shopping with us today!'
                # create a top border
                print('#' * 50)
                # print company information first using format
                print('\t\t{}'.format(company_name.title()))
                print('\t\t{}'.format(company_address.title()))
                print('\t\t{}'.format(company_city.title()))                
                # create a top border
                print('-' * 50)
                # print out header for section of items
                print('\tProduct Name\t   Quantity\tPrice')
                # create a print statement for each item
                for i in cart:
                    print('\t{}\t-{}-\t{} AUD'.format(i['med_name'],i['med_qty'],i['med_price']*i['med_qty']))                    
                # print a line between sections
                print('=' * 50)
                # print out header for section of total
                print('\t\t\tTotal')
                # calculate total price and print out
                for i in cart:
                    total = total + (float(i['med_price'])*i['med_qty'])
                print('\t\t\t${}'.format(total))
                # print a line between sections
                print('=' * 50)
                # output thank you message
                print('\n\t{}\n'.format(message))
                # create a bottom border
                print('-' * 50)

                #Creating an Invoice file in project folder
                now = datetime.now()
                print(f"Invoice Date ={now}\n")
                print('#' * 50)
                dt_string = now.strftime("%d%m%Y%H%M%S")                
                invoice_file_name = 'invoice_'+dt_string
                with  open(invoice_file_name,'w') as f:
                    # create a top border
                    f.write('#' * 50)
                    # print company information first using format
                    f.write('\n\t\t{}'.format(company_name.title()))
                    f.write('\n\t\t{}'.format(company_address.title()))
                    f.write('\n\t\t{}'.format(company_city.title()))                
                    # create a top border
                    f.write('\n')
                    f.write('-' * 50)
                    # print out header for section of items
                    f.write('\n\tProduct Name\tQuantity\tPrice\n')                    
                    # create a print statement for each item
                    for i in cart:
                        f.write('\t{}\t-{}-\t{} AUD'.format(i['med_name'],i['med_qty'],i['med_price']*i['med_qty']))
                        f.write('\n')
                    # print a line between sections
                    f.write('\n')
                    f.write('=' * 50)
                    # print out header for section of total
                    f.write('\n\t\t\tTotal')
                    # calculate total price and print out
                    f.write('\n\t\t\t${}'.format(total))
                    # print a line between sections
                    f.write('\n')
                    f.write('=' * 50)
                    # output thank you message
                    f.write('\n\t{}\n'.format(message))
                    # create a bottom border
                    f.write('\n')
                    f.write('-' * 50)
                    f.write(f"\nInvoice Date ={now}\n")
                    f.write('#' * 50)

                    #Reducing the inventory after the customer order has been processed
                    medicine_inventory = openpyxl.load_workbook("MEDICINE_INVENTORY.xlsx")
                    update_mi = medicine_inventory.active
                    for i in cart:
                        for J in range(2, update_mi.max_row+1):
                            update_med_id = update_mi.cell(row=J, column=1)
                            update_med_qty = update_mi.cell(row=J, column=4)
                            if (i["med_id"] == int(update_med_id.value)):
                                
                                #update the quantity
                                update_mi.cell(row=J, column=4).value = update_med_qty.value - i["med_qty"]
                    #saving the changes
                    medicine_inventory.save("MEDICINE_INVENTORY.xlsx")
                return None
        
            elif billing_user_input.lower() == "n":
                    print("Returning to the Cutomer Order Menu..")
                    time.sleep(1)
                    return None
            else:
                raise ValueError
        except ValueError:
            print("Invalid Option !")
            time.sleep(1)