
import re
import openpyxl
import time


def placing_customer_order(medicine_cart):
    found_med_flag = False
    qty_available_flag = False
    # Enter the med id
    while True:
        print("\n")
        print("****************************************************************")
        print("********************Placing Customer Order**********************")
        print("****************************************************************")
        #User input of the med id
        while True:
            try:
                med_id_user_input = int(input("Enter the med id to add to cart.. To discontinue, enter 0 : "))
                break
            except ValueError:
                print("Invalid med id !")
                time.sleep(2)
                continue
        #Return to Customer Order Menu    
        if med_id_user_input == 0:
            return medicine_cart
        #User input of med quantity
        while True:
            try:
                med_qty_user_input = int(input("Enter the quantity to add to cart.. To discontinue, enter 0 : "))
                break
            except ValueError:
                print("Invalid Quantity !")
                time.sleep(2)
                continue
        #Return to Customer Order Menu
        if med_qty_user_input == 0:
            return medicine_cart
        #Open inventory
        medicine_inventory = openpyxl.load_workbook("MEDICINE_INVENTORY.xlsx")
        mi = medicine_inventory.active
        # Search Inventory for the Medicine and quantity
        for i in range(2, mi.max_row+1):                            
            med_id = mi.cell(row=i, column=1)
            med_name = mi.cell(row=i, column=2)
            med_qty = mi.cell(row=i, column=4)
            if (med_id_user_input == int(med_id.value)):
                found_med_flag = True
                available_med_qty = med_qty.value
                if (med_qty_user_input <= int(med_qty.value)):
                    qty_available_flag = True                                   
                    if  (next((i for i, x in enumerate(medicine_cart) if x["med_id"] == med_id.value), None) is not None):
                        while True:
                            try:
                                med_duplicate_user_input = input("Med already present in cart. Do you want to replace the existing item in the cart? (y/n)")
                                if (med_duplicate_user_input.lower() =='y'):
                                    medicine_cart = list(filter(lambda i: i['med_id'] != med_id.value, medicine_cart))
                                    medicine_cart.append({'med_id': med_id.value,'med_name': med_name.value,'med_qty': med_qty_user_input})
                                    print("Medicine replaced !")
                                    break
                                elif (med_duplicate_user_input.lower() =='n'):
                                    print("Operation cancelled . Item not added to cart...")
                                    break
                                else:
                                    raise ValueError
                            except ValueError:
                                print("Invalid Entry (Only y/n) !")
                                time.sleep(1)
                                continue
                    else:
                        medicine_cart.append({'med_id': med_id.value,'med_name': med_name.value,'med_qty': med_qty_user_input})
                    print(f" MEDICINE CART : {medicine_cart}", end="\n")
        if found_med_flag is False:
            print("Medicine Not Found !")
        if found_med_flag is True and qty_available_flag is False:
            print(f"There is not enough quantity available for this medicine in the inventory. Only {available_med_qty} available !")
        
        
        
        
