import openpyxl

# Function for updating the invetory after an order is complete


def inventory_update(cart):
    # Reducing the inventory after the customer order has been processed
    # open the inventory
    medicine_inventory = openpyxl.load_workbook(
        "src/steffs_pharmacy/data/medicine_inventory.xlsx")
    update_mi = medicine_inventory.active
    for i in cart:
        for j in range(2, update_mi.max_row+1):
            update_med_id = update_mi.cell(row=j, column=1)
            update_med_qty = update_mi.cell(row=j, column=4)
            if (i["med_id"] == int(update_med_id.value)):
                # update the quantity
                update_mi.cell(
                    row=j, column=4).value = update_med_qty.value - i["med_qty"]
                # saving the changes
                medicine_inventory.save(
                    "src/steffs_pharmacy/data/medicine_inventory.xlsx")
