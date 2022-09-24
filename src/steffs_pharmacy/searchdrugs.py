import time
import re
import openpyxl
import clearing
from rich.console import Console
from rich.table import Table

# Function to search for Medicines


def search_drugs():
    clearing.clear()
    console = Console()
    table = Table(show_header=False, header_style="bold blue",
                  title="SEARCH FOR MEDICINE", title_justify="center")
    table.add_row("Search your inventory for medicines ..")
    # display header
    console.print(table)
    while (True):
        found = False
        # User input for drugs
        drug_name = input("\nEnter the drug name to search : ")
        print(f'Searching for {drug_name} in the inventory...\n')
        # Open the inventory file
        medicine_inventory = openpyxl.load_workbook(
            "src/steffs_pharmacy/data/medicine_inventory.xlsx")
        mi = medicine_inventory.active
        for i in range(1, mi.max_row+1):
            for j in range(1, 2):
                med_id = mi.cell(row=i, column=1)
                med_name = mi.cell(row=i, column=2)
                med_qty = mi.cell(row=i, column=4)
                med_price = mi.cell(row=i, column=6)
                # Search for medicine in the inventory
                match_drug = re.match(
                    '.*'+drug_name.lower()+'.*', med_name.value.lower())
                if (match_drug):
                    # medicine is found in inventory then display the same
                    found = True
                    print(f" Med Id : {med_id.value}", end=" ")
                    print(f" Med Name : {med_name.value}", end=" ")
                    print(f" Qty Available : {med_qty.value}", end="\n")
                    print(f" Price : {med_price.value}", end="\n")
        if found is False:
            # if medicine not found then display the alert
            print("Medicine Not Found !\n")
        while (True):
            try:
                # user input to  check if search need to be continued
                continue_search = str(
                    input("\nDo you wish to search for another medicine? (y/n): "))
                if continue_search == 'y':
                    break
                # Return to called function if 'n'
                elif continue_search == 'n':
                    return None
                else:
                    # Show alert if user input is invalid
                    print("Invalid Option ! Enter y/n ..")
            except ValueError:
                # show alert if user input is invalid
                print("Invalid Option !")
