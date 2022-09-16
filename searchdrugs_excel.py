import time
import re
import openpyxl

def searchdrugs():
    found =False
    print("****************************************************************")
    print("***************************Search Drugs*********************************")
    print("****************************************************************")
    drug_name=input("Enter the drug name to search : ")
    print(f'Searching for {drug_name} in the inventory...' )    
    #Open the inventory file
    medicine_inventory = openpyxl.load_workbook("MEDICINE_INVENTORY.xlsx")
    mi = medicine_inventory.active

    for i in range(1,mi.max_row+1):   
      
        for j in range(1,2):
            cell_object = mi.cell(row=i, column=1)
            # print(cell_object.value,end=" ")    
            # print("\n") 
            match_drug = re.match('.*'+drug_name.lower()+'.*',cell_object.value.lower())
            if (match_drug):                
                print(cell_object.value)
                found = True
    if found is False:
        print("Medicine Not Found")

searchdrugs()